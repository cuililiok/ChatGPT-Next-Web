#!/usr/bin/env python3
"""
达摩达兰DCF估值引擎 (Damodaran DCF Valuation Engine)

基于 Aswath Damodaran 的 fcffsimpleginzu 模板逻辑，
用 Python 实现完整的三阶段 FCFF DCF 估值框架。

核心设计原则（来自达摩达兰）：
1. Growth is value creating only when ROIC > WACC
2. Terminal value typically accounts for 60-80% of total value
3. Use normalized earnings for cyclical companies
4. Separate maintenance capex from growth capex

用法:
    from damodaran_dcf import DamodaranDCF

    model = DamodaranDCF()
    model.set_base_data(revenue=2032, ebit=166, ...)
    model.set_growth_assumptions(g1=10, g2=5, g_terminal=3, ...)
    model.set_cost_of_capital(rf=2.0, beta=1.1, erp=7.0, ...)
    result = model.run()
    result.print_summary()
    result.to_excel("output.xlsx")
"""

import os
import json
import copy
from dataclasses import dataclass, field
from typing import Optional
import statistics

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


@dataclass
class BaseData:
    """基础年份数据（Input Sheet）"""
    company_name: str = "Company"
    valuation_date: str = ""
    currency: str = "CNY (millions)"

    # 基础财务数据
    revenue: float = 0.0           # 营业收入
    operating_income: float = 0.0  # EBIT / 营业利润
    interest_expense: float = 0.0  # 利息支出
    book_equity: float = 0.0       # 账面净资产
    book_debt: float = 0.0         # 账面有息负债
    cash: float = 0.0              # 现金及等价物
    non_operating_assets: float = 0.0  # 非经营性资产（少数股东权益等）
    minority_interests: float = 0.0    # 少数股东权益
    shares_outstanding: float = 0.0    # 总股本（亿股）
    stock_price: float = 0.0           # 当前股价
    effective_tax_rate: float = 0.0    # 有效税率
    marginal_tax_rate: float = 0.0     # 边际税率

    # 员工期权
    has_options: bool = False
    options_outstanding: float = 0.0
    option_avg_strike: float = 0.0
    option_avg_maturity: float = 0.0

    # R&D 资本化
    has_rd: bool = False

    # 经营租赁
    has_operating_leases: bool = False

    # 折旧摊销（用于FCFF计算）
    depreciation_amortization: float = 0.0

    # 资本开支
    capex: float = 0.0


@dataclass
class GrowthAssumptions:
    """增长假设（Value Drivers）"""
    # 第一阶段 (Year 1)
    revenue_growth_y1: float = 0.0       # 下一年收入增长率
    operating_margin_y1: float = 0.0     # 下一年营业利润率

    # 第二阶段 (Years 2-5)
    revenue_growth_y2_5: float = 0.0     # 2-5年复合收入增长率
    target_operating_margin: float = 0.0  # 目标营业利润率
    year_of_margin_convergence: int = 5   # 利润率收敛年份

    # 第三阶段 (Years 6-10)
    sales_to_capital_y1_5: float = 0.0   # 前5年收入/资本比率
    sales_to_capital_y6_10: float = 0.0  # 后5年收入/资本比率

    # 终值
    terminal_growth_rate: float = 0.03   # 永续增长率
    terminal_value_method: str = "gordon"  # 终值方法: "gordon"(默认), "negative_growth"(负增长永续), "growing_annuity"(增长年金)
    growing_annuity_years: int = 15     # 增长年金模型额外增长年数（默认15年）


@dataclass
class CostOfCapital:
    """资本成本参数"""
    risk_free_rate: float = 0.0      # 无风险利率
    beta: float = 0.0                # 权益Beta
    equity_risk_premium: float = 0.0 # 市场风险溢价(ERP)
    country_risk_premium: float = 0.0 # 国家风险溢价(CRP)

    # 负债
    pre_tax_cost_of_debt: float = 0.0  # 税前负债成本
    debt_rating: str = ""              # 债务评级（可选）

    # 计算结果
    cost_of_equity: float = 0.0       # 权益成本 Ke
    cost_of_debt_after_tax: float = 0.0  # 税后负债成本
    wacc: float = 0.0                 # 加权平均资本成本
    equity_weight: float = 0.0        # 权益权重（市值）
    debt_weight: float = 0.0          # 负债权重（市值）

    # 终值WACC（收敛到行业均值）
    terminal_cost_of_capital: float = 0.0


@dataclass
class EPVResult:
    """EPV（盈利能力价值）结果"""
    normalized_ebit: float = 0.0
    normalized_tax_rate: float = 0.0
    nopat: float = 0.0
    maintenance_da: float = 0.0
    maintenance_capex: float = 0.0
    maintenance_fcff: float = 0.0
    wacc: float = 0.0
    epv_enterprise_value: float = 0.0
    net_debt: float = 0.0
    epv_equity_value: float = 0.0
    epv_per_share: float = 0.0
    book_value_per_share: float = 0.0
    epv_to_book_ratio: float = 0.0


@dataclass
class DCFYear:
    """DCF单年数据"""
    year: int = 0
    revenue_growth: float = 0.0
    revenue: float = 0.0
    operating_margin: float = 0.0
    ebit: float = 0.0
    ebit_1t: float = 0.0
    reinvestment: float = 0.0
    fcff: float = 0.0
    cost_of_capital: float = 0.0
    pv_fcff: float = 0.0
    invested_capital: float = 0.0
    roic: float = 0.0
    sales_to_capital: float = 0.0


@dataclass
class DCFResult:
    """DCF估值结果"""
    years: list = field(default_factory=list)
    pv_fcff_total: float = 0.0
    terminal_value: float = 0.0
    pv_terminal_value: float = 0.0
    enterprise_value: float = 0.0
    net_debt: float = 0.0
    equity_value: float = 0.0
    value_per_share: float = 0.0
    stock_price: float = 0.0
    price_to_value_ratio: float = 0.0  # 股价/价值
    safety_margin: float = 0.0        # 安全边际

    # 终值占比
    terminal_value_pct: float = 0.0

    # 诊断信息
    revenue_y10: float = 0.0
    ebit_y10: float = 0.0
    roic_y10: float = 0.0
    marginal_roic: float = 0.0


@dataclass
class SensitivityResult:
    """敏感性分析结果"""
    g_values: list = field(default_factory=list)
    margin_values: list = field(default_factory=list)
    per_share_matrix: list = field(default_factory=list)


class DamodaranDCF:
    """达摩达兰DCF估值引擎"""

    def __init__(self):
        self.base = BaseData()
        self.growth = GrowthAssumptions()
        self.coc = CostOfCapital()

    # ========== 数据设置方法 ==========

    def set_base_data(self, **kwargs):
        """设置基础年份数据"""
        for k, v in kwargs.items():
            if hasattr(self.base, k):
                setattr(self.base, k, v)
        return self

    def set_growth_assumptions(self, **kwargs):
        """设置增长假设"""
        for k, v in kwargs.items():
            if hasattr(self.growth, k):
                setattr(self.growth, k, v)
        return self

    def set_cost_of_capital(self, **kwargs):
        """设置资本成本参数"""
        for k, v in kwargs.items():
            if hasattr(self.coc, k):
                setattr(self.coc, k, v)
        return self

    # ========== EPV计算 ==========

    def compute_epv(self, normalized_ebit=None, normalized_tax_rate=None,
                    maintenance_da=None, maintenance_capex=None, wacc=None):
        """
        计算EPV（盈利能力价值）——假设零增长、只维持当前规模。

        这是达摩达兰估值框架的第一层：不假设任何成长，
        只评估公司"维持现状"值多少钱。

        参数:
            normalized_ebit: 正常化EBIT（若为None则用base.operating_income）
            normalized_tax_rate: 正常化税率（若为None则用base.effective_tax_rate）
            maintenance_da: 维持性折旧摊销（若为None则用base.depreciation_amortization）
            maintenance_capex: 维持性资本开支（若为None则用maintenance_da*1.05）
            wacc: 折现率（若为None则用计算出的coc.wacc）

        返回: EPVResult
        """
        ebit = normalized_ebit or self.base.operating_income
        tax_rate = normalized_tax_rate or self.base.effective_tax_rate
        da = maintenance_da or self.base.depreciation_amortization
        capex = maintenance_capex or (da * 1.05 if da > 0 else 0.0)
        w = wacc or self.coc.wacc

        if w <= 0:
            raise ValueError("WACC must be positive. Run compute_cost_of_capital first.")

        nopat = ebit * (1 - tax_rate)
        fcff = nopat + da - capex
        ev = fcff / w
        net_debt = self.base.book_debt - self.base.cash
        equity = ev - net_debt + self.base.non_operating_assets - self.base.minority_interests
        per_share = equity / self.base.shares_outstanding if self.base.shares_outstanding > 0 else 0
        book_ps = self.base.book_equity / self.base.shares_outstanding if self.base.shares_outstanding > 0 else 0

        return EPVResult(
            normalized_ebit=ebit,
            normalized_tax_rate=tax_rate,
            nopat=nopat,
            maintenance_da=da,
            maintenance_capex=capex,
            maintenance_fcff=fcff,
            wacc=w,
            epv_enterprise_value=ev,
            net_debt=net_debt,
            epv_equity_value=equity,
            epv_per_share=per_share,
            book_value_per_share=book_ps,
            epv_to_book_ratio=per_share / book_ps if book_ps > 0 else 0,
        )

    # ========== 基期正规化（周期性公司专用） ==========

    def normalize_base_data(self, normalized_ebit=None, normalized_margin=None,
                           normalized_capex_ratio=None, normalized_da_ratio=None,
                           historical_avg_margin=None, current_revenue=None):
        """
        对基期数据进行全参数正规化（周期性公司专用）。

        基于达摩达兰第3版第13章：正规化必须覆盖全参数——
        不仅盈利要正规化，再投资率和融资成本也需要正规化。

        使用方式（三选一，按优先级）：
        1. 直接传入 normalized_ebit → 替换 base.operating_income
        2. 传入 normalized_margin → 用 margin × revenue 计算
        3. 传入 historical_avg_margin → 自动计算

        参数:
            normalized_ebit: 正常化EBIT绝对值（直接覆盖）
            normalized_margin: 正常化EBIT利润率（与当期收入相乘）
            normalized_capex_ratio: 正常化CapEx/Revenue比率（默认用当期比率）
            normalized_da_ratio: 正常化D&A/Revenue比率（默认用当期比率）
            historical_avg_margin: 历史平均EBIT利润率（自动计算normalied_ebit）
            current_revenue: 当期收入（仅当传入historical_avg_margin时需要）

        返回: dict, 包含所有正规化后的值，供检查和审计
        """
        b = self.base
        result = {"原始EBIT": b.operating_income, "原始CapEx": b.capex,
                  "原始D&A": b.depreciation_amortization}

        # 步骤1：正规化EBIT
        if normalized_ebit is not None:
            b.operating_income = normalized_ebit
            result["EBIT来源"] = "直接传入"
        elif normalized_margin is not None:
            b.operating_income = normalized_margin * b.revenue
            result["EBIT来源"] = f"正常化利润率{normalized_margin*100:.1f}% × 收入{b.revenue}"
        elif historical_avg_margin is not None:
            rev = current_revenue or b.revenue
            b.operating_income = historical_avg_margin * rev
            result["EBIT来源"] = f"历史平均利润率{historical_avg_margin*100:.1f}% × 收入{rev}"
        else:
            result["EBIT来源"] = "未正规化（使用当期EBIT）"

        # 步骤2：正规化CapEx（如果有正常化比率传入）
        if normalized_capex_ratio is not None:
            old_capex = b.capex
            b.capex = normalized_capex_ratio * b.revenue
            result["正常化CapEx"] = b.capex
            result["CapEx来源"] = f"正常化比率{normalized_capex_ratio*100:.1f}% × 收入"
            if old_capex > 0:
                result["CapEx变化"] = f"{(b.capex/old_capex-1)*100:+.1f}%"
        else:
            result["正常化CapEx"] = b.capex
            result["CapEx来源"] = "未正规化"

        # 步骤3：正规化D&A（如果有正常化比率传入）
        if normalized_da_ratio is not None:
            old_da = b.depreciation_amortization
            b.depreciation_amortization = normalized_da_ratio * b.revenue
            result["正常化D&A"] = b.depreciation_amortization
            result["D&A来源"] = f"正常化比率{normalized_da_ratio*100:.1f}% × 收入"
        else:
            result["正常化D&A"] = b.depreciation_amortization
            result["D&A来源"] = "未正规化"

        result["正常化后EBIT"] = b.operating_income
        return result

    def compute_cost_of_capital_industry_beta(self, industry_unlevered_beta=None,
                                               target_de_ratio=None, erp=None):
        """
        使用行业Unlevered Beta替代个股回归Beta（A股推荐）。

        A股市场Beta噪音极大（个人投资者占比高、涨跌停板），
        Damodaran推荐使用行业Unlevered Beta重新加杠杆。

        参数:
            industry_unlevered_beta: 行业无杠杆Beta
            target_de_ratio: 目标D/E比率（若为None用当前市值比率）
            erp: 股权风险溢价（若为None用当前设置）

        公式:
            beta_levered = beta_unlevered × [1 + (1-t) × D/E]
        """
        c = self.coc
        b = self.base
        if industry_unlevered_beta is None:
            raise ValueError("必须提供行业无杠杆Beta")

        tax_rate = b.effective_tax_rate if b.effective_tax_rate > 0 else b.marginal_tax_rate
        de = target_de_ratio if target_de_ratio is not None else (
            b.book_debt / b.book_equity if b.book_equity > 0 else 0.5)

        c.beta = industry_unlevered_beta * (1 + (1 - tax_rate) * de)

        # 重新计算WACC
        return self.compute_cost_of_capital()

    # ========== WACC计算 ==========

    def compute_cost_of_capital(self):
        """计算WACC（达摩达兰方法）"""
        b = self.base
        c = self.coc

        # 权益成本: CAPM
        c.cost_of_equity = c.risk_free_rate + c.beta * (c.equity_risk_premium + c.country_risk_premium)

        # 税后负债成本
        tax_rate = b.effective_tax_rate if b.effective_tax_rate > 0 else b.marginal_tax_rate
        c.cost_of_debt_after_tax = c.pre_tax_cost_of_debt * (1 - tax_rate)

        # 权重：用市值
        mkt_cap = b.shares_outstanding * b.stock_price if (b.shares_outstanding > 0 and b.stock_price > 0) else b.book_equity
        net_debt = b.book_debt - b.cash
        total_value = mkt_cap + max(net_debt, 0)

        if total_value > 0:
            c.equity_weight = mkt_cap / total_value
            c.debt_weight = max(net_debt, 0) / total_value
        else:
            c.equity_weight = 0.8
            c.debt_weight = 0.2

        # WACC
        c.wacc = c.equity_weight * c.cost_of_equity + c.debt_weight * c.cost_of_debt_after_tax

        # 终值WACC（假设收敛到行业均值，默认用当前WACC）
        c.terminal_cost_of_capital = max(c.wacc - 0.005, self.growth.terminal_growth_rate + 0.05)

        return c

    # ========== DCF主计算 ==========

    def run(self, years=10):
        """
        执行完整的达摩达兰三阶段DCF估值。

        阶段划分:
        - Year 1: 使用 revenue_growth_y1 和 operating_margin_y1
        - Years 2-5: 使用 revenue_growth_y2_5，利润率线性收敛至target
        - Years 6-10: 利润率保持target，增长率收敛至terminal_growth_rate

        返回: DCFResult
        """
        b = self.base
        g = self.growth
        c = self.coc

        if c.wacc <= 0:
            self.compute_cost_of_capital()

        tax_rate = b.effective_tax_rate if b.effective_tax_rate > 0 else b.marginal_tax_rate
        base_margin = b.operating_income / b.revenue if b.revenue > 0 else 0

        # 初始值回填（仅当用户未显式设置时才回填，用 is None 或 ==0 判断）
        # 注意：不能用 `or`，否则合法的 0 值（如亏损公司利润率=0）会被错误替换
        if g.operating_margin_y1 == 0:
            g.operating_margin_y1 = base_margin
        if g.target_operating_margin == 0:
            g.target_operating_margin = g.operating_margin_y1
        fallback_stc = b.revenue / max(b.book_equity + b.book_debt, 1) if b.revenue > 0 else 1.0
        if g.sales_to_capital_y1_5 == 0:
            g.sales_to_capital_y1_5 = fallback_stc
        if g.sales_to_capital_y6_10 == 0:
            g.sales_to_capital_y6_10 = g.sales_to_capital_y1_5

        invested_capital = b.book_equity + b.book_debt  # 初始投入资本

        dcf_years = []
        revenue = b.revenue
        pv_fcff_sum = 0.0
        cum_discount = 1.0
        final_cum_discount = 1.0  # 保存第10年末的累积折现因子，用于终值折现

        for yr in range(1, years + 1):
            dy = DCFYear(year=yr)

            # 增长率
            if yr == 1:
                dy.revenue_growth = g.revenue_growth_y1
            elif yr <= 5:
                dy.revenue_growth = g.revenue_growth_y2_5
            else:
                # Years 6-10: 线性收敛至终值增长率
                remaining = years - 5
                progress = (yr - 5) / remaining
                dy.revenue_growth = g.revenue_growth_y2_5 + progress * (g.terminal_growth_rate - g.revenue_growth_y2_5)

            # 收入
            revenue = revenue * (1 + dy.revenue_growth)
            dy.revenue = revenue

            # 利润率（线性收敛）
            if yr == 1:
                dy.operating_margin = g.operating_margin_y1
            elif yr <= g.year_of_margin_convergence:
                progress = (yr - 1) / (g.year_of_margin_convergence - 1) if g.year_of_margin_convergence > 1 else 1.0
                dy.operating_margin = g.operating_margin_y1 + progress * (g.target_operating_margin - g.operating_margin_y1)
            else:
                dy.operating_margin = g.target_operating_margin

            # EBIT
            dy.ebit = revenue * dy.operating_margin
            dy.ebit_1t = dy.ebit * (1 - tax_rate)

            # 再投资 = 收入增长额 / Sales_to_Capital
    # Sales_to_Capital = Revenue / Invested Capital
    # 含义：每1元资本能产生多少元收入
            delta_rev = revenue - (revenue / (1 + dy.revenue_growth))
            stc = g.sales_to_capital_y1_5 if yr <= 5 else g.sales_to_capital_y6_10
            dy.reinvestment = delta_rev / stc if stc > 0 else 0
            dy.sales_to_capital = stc

            # FCFF
            dy.fcff = dy.ebit_1t - dy.reinvestment

            # 折现
            # WACC从初始值线性收敛至终值WACC（Years 1-10）
            dy.cost_of_capital = c.wacc + (yr / years) * (c.terminal_cost_of_capital - c.wacc)
            cum_discount *= 1 / (1 + dy.cost_of_capital)
            dy.pv_fcff = dy.fcff * cum_discount
            pv_fcff_sum += dy.pv_fcff
            final_cum_discount = cum_discount  # 每年更新，循环结束后即为第N年末的累积折现因子

            # 投入资本和ROIC
            invested_capital += dy.reinvestment
            dy.invested_capital = invested_capital
            dy.roic = dy.ebit * (1 - tax_rate) / invested_capital if invested_capital > 0 else 0

            dcf_years.append(dy)

        # 终值（支持三种方法：Gordon Growth / 负增长永续 / 增长年金）
        last_year = dcf_years[-1]
        terminal_fcff = last_year.ebit_1t - last_year.reinvestment
        g_term = g.terminal_growth_rate
        w_terminal = c.terminal_cost_of_capital
        method = g.terminal_value_method

        if method == "growing_annuity":
            # 方法：增长年金（有限增长期，不假设永续）
            # TV = sum of FCFF from year N+1 to N+T, discounted back to year N
            # Then discount the year-N value to present using final_cum_discount
            T = g.growing_annuity_years
            tv_at_year_n = 0.0
            for t in range(1, T + 1):
                # Assume FCFF grows at terminal_growth_rate for each extra year
                fcff_t = terminal_fcff * (1 + g_term) ** t
                tv_at_year_n += fcff_t / (1 + w_terminal) ** t
            terminal_value = tv_at_year_n
            pv_terminal = terminal_value * final_cum_discount

        elif method == "negative_growth":
            # 方法：负增长永续（适用于结构性萎缩公司）
            # TV = FCFF_{N+1} / (WACC - g_negative), where g_negative < 0
            # 约束：|g_negative| < WACC (否则TV为负)
            if abs(g_term) >= w_terminal:
                g_term = -(w_terminal - 0.01)  # 钳制，确保TV为正
            if terminal_fcff > 0 and w_terminal + abs(g_term) > 0:
                terminal_value = terminal_fcff * (1 + g_term) / (w_terminal - g_term)
                pv_terminal = terminal_value * final_cum_discount
            else:
                terminal_value = 0
                pv_terminal = 0

        else:
            # 默认方法：Gordon Growth Model
            # 终值折现：使用与FCFF相同的累积折现因子（final_cum_discount），
            # 而非 (1+WACC)^n，因为WACC在预测期内是随时间收敛的变动值，
            # 两者必须使用同一折现路径，否则终值会被系统性高估。
            if terminal_fcff > 0 and w_terminal > g_term:
                terminal_value = terminal_fcff * (1 + g_term) / (w_terminal - g_term)
                pv_terminal = terminal_value * final_cum_discount
            elif w_terminal <= g_term:
                w_terminal = g_term + 0.02
                terminal_value = terminal_fcff * (1 + g_term) / (w_terminal - g_term)
                pv_terminal = terminal_value * final_cum_discount
            else:
                terminal_value = 0
                pv_terminal = 0

        enterprise_value = pv_fcff_sum + pv_terminal

        # 股权价值
        net_debt = b.book_debt - b.cash
        equity_value = enterprise_value - net_debt + b.non_operating_assets - b.minority_interests
        value_per_share = equity_value / b.shares_outstanding if b.shares_outstanding > 0 else 0

        # 安全边际
        if value_per_share > 0:
            safety_margin = (value_per_share - b.stock_price) / value_per_share
        else:
            safety_margin = 0

        # 终值占比
        tv_pct = pv_terminal / enterprise_value if enterprise_value > 0 else 0

        # 边际ROIC（Years 1-10平均）
        marginal_roics = [dy.roic for dy in dcf_years[1:]]
        marginal_roic = statistics.mean(marginal_roics) if marginal_roics else 0

        result = DCFResult(
            years=dcf_years,
            pv_fcff_total=pv_fcff_sum,
            terminal_value=terminal_value,
            pv_terminal_value=pv_terminal,
            enterprise_value=enterprise_value,
            net_debt=net_debt,
            equity_value=equity_value,
            value_per_share=value_per_share,
            stock_price=b.stock_price,
            price_to_value_ratio=b.stock_price / value_per_share if value_per_share > 0 else 999,
            safety_margin=safety_margin,
            terminal_value_pct=tv_pct,
            revenue_y10=last_year.revenue,
            ebit_y10=last_year.ebit,
            roic_y10=last_year.roic,
            marginal_roic=marginal_roic,
        )

        self._last_result = result
        return result

    # ========== 情景分析 ==========

    def run_scenarios(self, scenarios: dict, probabilities: dict = None) -> dict:
        """
        运行多情景DCF估值，支持概率加权（达摩达兰第3版第3章+第13章推荐方法）。

        参数:
            scenarios: dict，key=情景名，value=dict of kwargs for set_growth_assumptions / set_cost_of_capital
            probabilities: dict，key=情景名，value=概率权重（0-1），概率之和应为1。
                          若为None，等权处理。
            示例: {
                "悲观": {"revenue_growth_y1": 5, "revenue_growth_y2_5": 3, "terminal_growth_rate": 2.0, "wacc_adj": 1.0},
                "基准": {"revenue_growth_y1": 10, ...},
                "乐观": {"revenue_growth_y1": 15, ...},
            }
            probabilities: {"悲观": 0.25, "基准": 0.50, "乐观": 0.25}

        返回: dict，key=情景名，value=DCFResult；若提供probabilities则额外包含"_summary"键
        """
        results = {}
        original_growth = {k: getattr(self.growth, k) for k in vars(self.growth)}
        original_coc = {k: getattr(self.coc, k) for k in vars(self.coc)}

        for name, params in scenarios.items():
            # 重置
            for k, v in original_growth.items():
                setattr(self.growth, k, v)
            for k, v in original_coc.items():
                setattr(self.coc, k, v)

            # 应用情景参数
            growth_params = {k: v for k, v in params.items() if hasattr(self.growth, k)}
            coc_params = {k: v for k, v in params.items() if hasattr(self.coc, k) or k == "wacc_adj"}

            if growth_params:
                self.set_growth_assumptions(**growth_params)
            if "wacc_adj" in coc_params:
                self.coc.wacc += coc_params.pop("wacc_adj")
            # Terminal CoC
            self.coc.terminal_cost_of_capital = max(self.coc.wacc - 0.01, self.growth.terminal_growth_rate + 0.05)

            result = self.run()
            results[name] = result

        # 恢复
        for k, v in original_growth.items():
            setattr(self.growth, k, v)
        for k, v in original_coc.items():
            setattr(self.coc, k, v)

        # 概率加权
        if probabilities:
            weighted_value = 0.0
            weighted_ev = 0.0
            prob_sum = sum(probabilities.get(name, 0) for name in scenarios)
            if abs(prob_sum - 1.0) > 0.01:
                print(f"[WARN] 概率之和={prob_sum:.2f}，已自动归一化")
                normalized_probs = {k: v / prob_sum for k, v in probabilities.items()}
            else:
                normalized_probs = probabilities

            for name, res in results.items():
                p = normalized_probs.get(name, 0)
                weighted_value += p * res.value_per_share
                weighted_ev += p * res.enterprise_value
                res.weighted_contribution = p * res.value_per_share
                res.probability = p

            all_values = [r.value_per_share for r in results.values() if hasattr(r, 'value_per_share')]
            results["_summary"] = {
                "weighted_value_per_share": weighted_value,
                "weighted_enterprise_value": weighted_ev,
                "var_5pct": min(all_values) if all_values else 0,
                "range": [min(all_values), max(all_values)] if all_values else [0, 0],
                "probabilities": {k: normalized_probs.get(k, 0) for k in scenarios},
            }

        return results

    # ========== 敏感性分析 ==========

    def sensitivity_analysis(self, g_range=None, margin_range=None, years=10):
        """
        终值增长率和利润率的敏感性分析。

        参数:
            g_range: 终值增长率列表 [2.0, 2.5, 3.0, 3.5, 4.0]
            margin_range: 终值利润率列表 [0.08, 0.09, 0.10, 0.11, 0.12]
        """
        if g_range is None:
            g_range = [0.02, 0.025, 0.03, 0.035, 0.04]
        if margin_range is None:
            margin_range = [0.08, 0.09, 0.10, 0.11, 0.12]

        original_target_margin = self.growth.target_operating_margin
        original_g = self.growth.terminal_growth_rate

        matrix = []
        for g in g_range:
            row = []
            for m in margin_range:
                self.growth.target_operating_margin = m
                self.growth.terminal_growth_rate = g
                result = self.run(years=years)
                row.append(result.value_per_share)
            matrix.append(row)

        # 恢复
        self.growth.target_operating_margin = original_target_margin
        self.growth.terminal_growth_rate = original_g

        return SensitivityResult(
            g_values=g_range,
            margin_values=margin_range,
            per_share_matrix=matrix,
        )

    # ========== 输出 ==========

    def print_summary(self, result=None):
        """打印估值摘要（终端彩色输出）"""
        if result is None:
            result = self._last_result
        if result is None:
            result = self.run()

        b = self.base
        c = self.coc

        print(f"\n{'='*60}")
        print(f"  {b.company_name} — 达摩达兰DCF估值")
        print(f"{'='*60}")
        print(f"  股价: {b.stock_price:.2f}  |  每股价值: {result.value_per_share:.2f}")
        print(f"  安全边际: {result.safety_margin*100:.1f}%")
        print(f"  {'低估' if result.safety_margin > 0.2 else '合理' if result.safety_margin > -0.2 else '高估'}")
        print(f"{'='*60}")
        print(f"  WACC: {c.wacc*100:.1f}% (Ke={c.cost_of_equity*100:.1f}%, Kd={c.cost_of_debt_after_tax*100:.1f}%)")
        print(f"  终值占比: {result.terminal_value_pct*100:.1f}%")
        print(f"  ROIC Y10: {result.roic_y10*100:.1f}% vs WACC {c.wacc*100:.1f}%")
        print(f"{'='*60}\n")

    def to_excel(self, result=None, filepath="damodaran_dcf_output.xlsx"):
        """输出估值结果到Excel（达摩达兰风格）"""
        if result is None:
            result = self._last_result
        if result is None:
            result = self.run()

        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required. Install: pip install openpyxl")

        b = self.base
        g = self.growth
        c = self.coc

        wb = openpyxl.Workbook()

        # ===== Sheet 1: Input Sheet =====
        ws = wb.active
        ws.title = "Input Sheet"

        hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        bdr = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws.merge_cells('A1:D1')
        ws['A1'] = f"{b.company_name} — Damodaran DCF Valuation"
        ws['A1'].font = Font(bold=True, size=14, color="003366")

        r = 3
        ws.cell(row=r, column=1, value="Input Cell").font = hdr_font
        ws.cell(row=r, column=1).fill = hdr_fill
        ws.cell(row=r, column=2, value="Value").font = hdr_font
        ws.cell(row=r, column=2).fill = hdr_fill
        ws.cell(row=r, column=3, value="Notes").font = hdr_font
        ws.cell(row=r, column=3).fill = hdr_fill

        inputs = [
            ("Date of valuation", b.valuation_date, ""),
            ("Company name", b.company_name, ""),
            ("Currency", b.currency, ""),
            ("", "", ""),
            ("Revenues", b.revenue, ""),
            ("Operating income (EBIT)", b.operating_income, ""),
            ("Interest expense", b.interest_expense, ""),
            ("Book value of equity", b.book_equity, ""),
            ("Book value of debt", b.book_debt, ""),
            ("Cash & Marketable Securities", b.cash, ""),
            ("Non-operating assets", b.non_operating_assets, ""),
            ("Minority interests", b.minority_interests, ""),
            ("Number of shares", b.shares_outstanding, ""),
            ("Current stock price", b.stock_price, ""),
            ("Effective tax rate", b.effective_tax_rate, ""),
            ("Marginal tax rate", b.marginal_tax_rate, ""),
            ("", "", ""),
            ("Value Drivers", "", ""),
            ("Revenue growth - next year", g.revenue_growth_y1, ""),
            ("Operating margin - next year", g.operating_margin_y1, ""),
            ("Revenue growth - years 2-5", g.revenue_growth_y2_5, ""),
            ("Target operating margin", g.target_operating_margin, ""),
            ("Year of margin convergence", g.year_of_margin_convergence, ""),
            ("Sales to Capital (years 1-5)", g.sales_to_capital_y1_5, ""),
            ("Sales to Capital (years 6-10)", g.sales_to_capital_y6_10, ""),
            ("Terminal growth rate", g.terminal_growth_rate, ""),
            ("", "", ""),
            ("Cost of Capital", "", ""),
            ("Risk-free rate", c.risk_free_rate, ""),
            ("Beta", c.beta, ""),
            ("Equity risk premium", c.equity_risk_premium, ""),
            ("Country risk premium", c.country_risk_premium, ""),
            ("Cost of equity (Ke)", c.cost_of_equity, f"= {c.risk_free_rate} + {c.beta} x ({c.equity_risk_premium} + {c.country_risk_premium})"),
            ("Pre-tax cost of debt", c.pre_tax_cost_of_debt, ""),
            ("After-tax cost of debt", c.cost_of_debt_after_tax, f"= {c.pre_tax_cost_of_debt} x (1 - {b.effective_tax_rate})"),
            ("WACC", c.wacc, f"= {c.equity_weight:.1%} x {c.cost_of_equity:.1%} + {c.debt_weight:.1%} x {c.cost_of_debt_after_tax:.1%}"),
        ]

        for i, (label, val, note) in enumerate(inputs):
            row = r + 1 + i
            if label and val != "":
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=val)
                ws.cell(row=row, column=3, value=note).font = Font(italic=True, color="666666")

        # ===== Sheet 2: Valuation Output =====
        ws2 = wb.create_sheet("Valuation Output")

        ws2.cell(row=1, column=1, value="Year").font = hdr_font
        ws2.cell(row=1, column=1).fill = hdr_fill
        for col, header in enumerate(["Revenue Growth", "Revenue", "Operating Margin", "EBIT", "EBIT(1-t)", "Reinvestment", "FCFF", "Cost of Capital", "PV(FCFF)", "ROIC"], 2):
            ws2.cell(row=1, column=col, value=header).font = hdr_font
            ws2.cell(row=1, column=col).fill = hdr_fill

        for i, dy in enumerate(result.years):
            row = 2 + i
            ws2.cell(row=row, column=1, value=dy.year)
            ws2.cell(row=row, column=2, value=dy.revenue_growth).number_format = '0.0%'
            ws2.cell(row=row, column=3, value=dy.revenue).number_format = '#,##0.0'
            ws2.cell(row=row, column=4, value=dy.operating_margin).number_format = '0.0%'
            ws2.cell(row=row, column=5, value=dy.ebit).number_format = '#,##0.0'
            ws2.cell(row=row, column=6, value=dy.ebit_1t).number_format = '#,##0.0'
            ws2.cell(row=row, column=7, value=dy.reinvestment).number_format = '#,##0.0'
            ws2.cell(row=row, column=8, value=dy.fcff).number_format = '#,##0.0'
            ws2.cell(row=row, column=9, value=dy.cost_of_capital).number_format = '0.0%'
            ws2.cell(row=row, column=10, value=dy.pv_fcff).number_format = '#,##0.0'
            ws2.cell(row=row, column=11, value=dy.roic).number_format = '0.0%'

        # 估值汇总
        sr = len(result.years) + 4
        summary = [
            ("PV(FCFF over next 10 years)", result.pv_fcff_total),
            ("Terminal Value", result.terminal_value),
            ("PV(Terminal Value)", result.pv_terminal_value),
            ("", ""),
            ("Enterprise Value", result.enterprise_value),
            ("- Net Debt", result.net_debt),
            ("+ Non-operating Assets", b.non_operating_assets),
            ("- Minority Interests", b.minority_interests),
            ("", ""),
            ("Equity Value", result.equity_value),
            ("Shares Outstanding", b.shares_outstanding),
            ("Value per Share", result.value_per_share),
            ("", ""),
            ("Stock Price", b.stock_price),
            ("Price / Value", result.price_to_value_ratio),
            ("Safety Margin", f"{result.safety_margin*100:.1f}%"),
        ]
        for i, (label, val) in enumerate(summary):
            row = sr + i
            ws2.cell(row=row, column=1, value=label).font = Font(bold=True)
            if isinstance(val, (int, float)):
                ws2.cell(row=row, column=2, value=val).number_format = '#,##0.00'

        # 列宽
        ws2.column_dimensions['A'].width = 30
        for col in range(2, 12):
            ws2.column_dimensions[get_column_letter(col)].width = 14

        # ===== Sheet 3: EPV =====
        ws3 = wb.create_sheet("EPV")
        epv = self.compute_epv()
        ws3.merge_cells('A1:C1')
        ws3['A1'] = "EPV (Earning Power Value) — Zero Growth"
        ws3['A1'].font = Font(bold=True, size=13, color="003366")

        epv_data = [
            ("Normalized EBIT", epv.normalized_ebit),
            ("Tax Rate", epv.normalized_tax_rate),
            ("NOPAT", epv.nopat),
            ("+ Maintenance D&A", epv.maintenance_da),
            ("- Maintenance Capex", epv.maintenance_capex),
            ("= Maintenance FCFF", epv.maintenance_fcff),
            ("", ""),
            ("/ WACC", epv.wacc),
            ("= EPV (Enterprise Value)", epv.epv_enterprise_value),
            ("- Net Debt", epv.net_debt),
            ("= EPV (Equity Value)", epv.epv_equity_value),
            ("/ Shares", b.shares_outstanding),
            ("= EPV per Share", epv.epv_per_share),
            ("", ""),
            ("Book Value per Share", epv.book_value_per_share),
            ("EPV / Book Value", epv.epv_to_book_ratio),
        ]
        for i, (label, val) in enumerate(epv_data):
            ws3.cell(row=3+i, column=1, value=label).font = Font(bold=True)
            if isinstance(val, (int, float)):
                ws3.cell(row=3+i, column=2, value=val).number_format = '#,##0.00'

        ws3.column_dimensions['A'].width = 30

        # ===== Sheet 4: Sensitivity =====
        ws4 = wb.create_sheet("Sensitivity")
        ws4['A1'] = "Terminal Value Sensitivity"
        ws4['A1'].font = Font(bold=True, size=13, color="003366")

        sens = self.sensitivity_analysis()
        ws4.cell(row=3, column=1, value="g ↓ / margin →").font = Font(bold=True)
        for j, m in enumerate(sens.margin_values):
            ws4.cell(row=3, column=2+j, value=f"{m*100:.1f}%").font = Font(bold=True)
            ws4.cell(row=3, column=2+j).fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        for i, g in enumerate(sens.g_values):
            ws4.cell(row=4+i, column=1, value=f"{g*100:.1f}%").font = Font(bold=True)
            for j, ps in enumerate(sens.per_share_matrix[i]):
                cell = ws4.cell(row=4+i, column=2+j, value=round(ps, 2))
                cell.number_format = '0.00'
                if ps > b.stock_price * 1.2:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif ps > b.stock_price * 0.8:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb.save(filepath)
        return filepath

    # ========== 达摩达兰诊断 ==========

    def diagnostics(self, result=None):
        """
        达摩达兰诊断检查（基于Diagnostics sheet的逻辑）。

        检查项:
        1. 增长率是否合理（vs 行业均值）
        2. 第10年收入是否合理（TAM约束）
        3. 利润率是否合理（vs 行业均值）
        4. 再投资是否与增长一致
        5. ROIC vs WACC（成长是否创造价值）
        """
        if result is None:
            result = self._last_result
        if result is None:
            result = self.run()

        checks = []

        # Check 1: ROIC vs WACC
        if result.marginal_roic > self.coc.wacc:
            checks.append(("PASS", f"边际ROIC ({result.marginal_roic*100:.1f}%) > WACC ({self.coc.wacc*100:.1f}%) — 成长创造价值"))
        else:
            checks.append(("FAIL", f"边际ROIC ({result.marginal_roic*100:.1f}%) < WACC ({self.coc.wacc*100:.1f}%) — 成长毁灭价值！"))

        # Check 2: 终值占比
        if result.terminal_value_pct < 0.5:
            checks.append(("WARN", f"终值占比仅{result.terminal_value_pct*100:.0f}% — 检查是否低估了公司持续经营价值"))
        elif result.terminal_value_pct > 0.9:
            checks.append(("WARN", f"终值占比{result.terminal_value_pct*100:.0f}% — 估值高度依赖终值假设"))
        else:
            checks.append(("PASS", f"终值占比{result.terminal_value_pct*100:.0f}% — 在合理范围内(50-90%)"))

        # Check 3: 第10年利润率
        margin_y10 = result.years[-1].operating_margin if result.years else 0
        base_margin = self.base.operating_income / self.base.revenue if self.base.revenue > 0 else 0
        if abs(margin_y10 - base_margin) > 0.1:
            checks.append(("WARN", f"第10年利润率({margin_y10*100:.1f}%)与基准({base_margin*100:.1f}%)差异>10pct — 需要强理由"))
        else:
            checks.append(("PASS", f"利润率收敛合理: {base_margin*100:.1f}% → {margin_y10*100:.1f}%"))

        # Check 4: 安全边际
        if result.safety_margin > 0.3:
            checks.append(("PASS", f"安全边际{result.safety_margin*100:.1f}% > 30% — 有显著安全边际"))
        elif result.safety_margin > 0:
            checks.append(("OK", f"安全边际{result.safety_margin*100:.1f}% — 有一定安全边际"))
        else:
            checks.append(("FAIL", f"安全边际{result.safety_margin*100:.1f}% — 股价高于估值"))

        return checks


# ========== 行业数据查询 ==========

class DamodaranData:
    """
    从达摩达兰模板中读取行业数据。

    需要达摩达兰模板文件（已打包在 assets/damodaran-templates/ 中）。
    """

    def __init__(self, templates_dir=None):
        self.templates_dir = templates_dir or os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "assets", "damodaran-templates"
        )

    def get_industry_averages(self, region="US"):
        """
        读取行业平均数据。

        参数:
            region: "US" 或 "Global"

        返回: dict, key=行业名, value={"firms": N, "growth": x, "margin": x, "roc": x}
        """
        if not HAS_XLRD:
            raise ImportError("xlrd required for reading .xls files. Install: pip install xlrd")

        fname = "betas.xls" if region == "US" else None
        # 行业数据在多个文件中，主要从betas.xls和spearn.xls获取
        industries = {}

        for fn in ["betas.xls", "spearn.xls"]:
            path = os.path.join(self.templates_dir, fn)
            if not os.path.exists(path):
                continue
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            for r in range(1, ws.nrows):
                try:
                    name = ws.cell_value(r, 0)
                    if name and isinstance(name, str) and len(name) > 2:
                        if name not in industries:
                            industries[name] = {}
                        # 尝试读取其他列
                        for c in range(1, min(10, ws.ncols)):
                            v = ws.cell_value(r, c)
                            if isinstance(v, float) and v != '':
                                industries[name][f"col_{c}"] = v
                except:
                    pass
            wb.release_resources()

        return industries

    def get_erp(self):
        """读取月度ERP数据。"""
        path = os.path.join(self.templates_dir, "ERPbymonth.xlsx")
        if os.path.exists(path) and HAS_PANDAS:
            return pd.read_excel(path)
        return None

    def get_country_risk_premiums(self):
        """读取各国风险溢价。"""
        path = os.path.join(self.templates_dir, "ctrypremApr26.xlsx")
        if os.path.exists(path) and HAS_PANDAS:
            return pd.read_excel(path)
        return None


# ========== 快速入门示例 ==========

if __name__ == "__main__":
    # 万华化学示例
    model = DamodaranDCF()
    model.set_base_data(
        company_name="万华化学 (600309)",
        valuation_date="2026-04-21",
        revenue=2032.35,
        operating_income=166.71,
        interest_expense=25.5,
        book_equity=1201.29,
        book_debt=1090.0,
        cash=200.0,
        non_operating_assets=0,
        minority_interests=0,
        shares_outstanding=31.30,
        stock_price=90.68,
        effective_tax_rate=0.13,
        marginal_tax_rate=0.25,
        depreciation_amortization=105.8,
        capex=520.0,
    )
    model.set_growth_assumptions(
        revenue_growth_y1=0.10,
        operating_margin_y1=0.095,
        revenue_growth_y2_5=0.07,
        target_operating_margin=0.105,
        year_of_margin_convergence=5,
        sales_to_capital_y1_5=0.30,
        sales_to_capital_y6_10=0.35,
        terminal_growth_rate=0.03,
    )
    model.set_cost_of_capital(
        risk_free_rate=0.02,
        beta=1.10,
        equity_risk_premium=0.07,
        country_risk_premium=0.0,
        pre_tax_cost_of_debt=0.045,
    )
    model.compute_cost_of_capital()

    # EPV
    epv = model.compute_epv(normalized_ebit=203.2, maintenance_da=105.8)
    print(f"EPV per share: {epv.epv_per_share:.2f}")
    print(f"EPV/Book: {epv.epv_to_book_ratio:.2f}x")

    # DCF
    result = model.run()
    model.print_summary(result)

    # 诊断
    print("\nDiagnostics:")
    for status, msg in model.diagnostics(result):
        icon = {"PASS": "[OK]", "FAIL": "[!!]", "WARN": "[??]", "OK": "[ok]"}[status]
        print(f"  {icon} {msg}")

    # 输出Excel
    output_path = os.path.join(os.path.dirname(__file__), "..", "..", "output", "damodaran_dcf_output.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    model.to_excel(result, filepath=output_path)
    print(f"\nExcel saved: {output_path}")