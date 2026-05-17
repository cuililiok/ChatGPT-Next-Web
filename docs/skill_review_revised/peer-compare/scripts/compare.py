#!/usr/bin/env python3
"""
同业财务对比工具
用法：python compare.py --stocks 600519 000858 000568 [--year 2023] [--output report.xlsx]

输出：
  - 终端彩色对比表（基本面 + 估值 + 成长性 + 现金质量）
  - 可选：导出 Excel 对比报告

数据来源：AKShare（东方财富 / 新浪财经）
"""

import argparse, sys, os, subprocess, datetime
from pathlib import Path

# ── 自动安装依赖 ──────────────────────────────────────────
def ensure(pkg, import_name=None):
    try:
        __import__(import_name or pkg)
    except ImportError:
        print(f"  正在安装 {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg,
                               "--break-system-packages", "-q"])

ensure("akshare")
ensure("pandas")
ensure("openpyxl")
ensure("rich")

import akshare as ak
import pandas as pd
from rich.console import Console
from rich.table import Table
from rich import box
from rich.text import Text

console = Console()

# ── 指标定义 ──────────────────────────────────────────────
# (字段名, 显示名, 单位, 越大越好?, 格式)
METRICS = {
    "valuation": {
        "title": "💰 估值",
        "fields": [
            ("pe_ttm",      "PE(TTM)",   "倍",  False, ".1f"),
            ("pb",          "PB",        "倍",  False, ".2f"),
            ("ps_ttm",      "PS(TTM)",   "倍",  False, ".2f"),
            ("dividend_yield", "股息率", "%",   True,  ".2f"),
            ("market_cap",  "总市值",    "亿",  None,  ".0f"),
        ]
    },
    "fundamental": {
        "title": "📊 基本面",
        "fields": [
            ("roe",         "ROE",       "%",   True,  ".2f"),
            ("gross_margin","毛利率",    "%",   True,  ".2f"),
            ("net_margin",  "净利率",    "%",   True,  ".2f"),
            ("debt_ratio",  "资产负债率","%",   False, ".2f"),
            ("current_ratio","流动比率", "倍",  True,  ".2f"),
        ]
    },
    "growth": {
        "title": "🚀 成长性",
        "fields": [
            ("revenue_growth",    "营收增速",   "%", True, ".2f"),
            ("profit_growth",     "净利润增速", "%", True, ".2f"),
            ("roe_3y_avg",        "ROE(3年均)", "%", True, ".2f"),
            ("eps_growth",        "EPS增速",   "%", True, ".2f"),
        ]
    },
    "quality": {
        "title": "🏆 现金质量",
        "fields": [
            ("cfo_to_profit",     "经营现金/净利",  "倍",  True,  ".2f"),
            ("fcf_yield",         "自由现金收益率", "%",   True,  ".2f"),
            ("receivable_ratio",  "应收/营收",      "%",   False, ".2f"),
            ("gross_profit_yoy",  "毛利额增速",     "%",   True,  ".2f"),
        ]
    }
}


def get_stock_name(code: str) -> str:
    """获取股票简称"""
    try:
        df = ak.stock_individual_info_em(symbol=code)
        row = df[df["item"] == "股票简称"]
        if not row.empty:
            return row.iloc[0]["value"]
    except Exception:
        pass
    return code


def fetch_realtime_metrics(code: str) -> dict:
    """从东方财富获取实时估值数据"""
    data = {}
    try:
        df = ak.stock_individual_info_em(symbol=code)
        info = dict(zip(df["item"], df["value"]))
        # 字段映射（东方财富返回的字段名可能因版本变化）
        def safe_float(v):
            try: return float(str(v).replace(",","").replace("%","").strip())
            except: return None

        data["market_cap"]    = safe_float(info.get("总市值"))
        if data["market_cap"]:
            data["market_cap"] /= 1e8   # 转亿元
    except Exception as e:
        pass

    try:
        # 实时行情（含PE/PB）
        df2 = ak.stock_a_pe_and_pb(symbol=code)  # 部分版本支持
        if not df2.empty:
            row = df2.iloc[-1]
            data.setdefault("pe_ttm", float(row.get("pe_ttm", 0)) or None)
            data.setdefault("pb",     float(row.get("pb", 0)) or None)
    except Exception:
        pass

    try:
        # 用 stock_zh_a_spot_em 拿实时行情
        df3 = ak.stock_zh_a_spot_em()
        row = df3[df3["代码"] == code]
        if not row.empty:
            r = row.iloc[0]
            def sf(col):
                try: return float(r[col])
                except: return None
            data.setdefault("pe_ttm",    sf("市盈率-动态"))
            data.setdefault("pb",        sf("市净率"))
            data.setdefault("market_cap", sf("总市值"))
            if data.get("market_cap"):
                data["market_cap"] /= 1e8
    except Exception:
        pass

    return data


def fetch_financial_metrics(code: str, year: str) -> dict:
    """从财报获取基本面数据"""
    data = {}

    # ── 利润表 ──
    try:
        df = ak.stock_profit_sheet_by_yearly_em(symbol=code)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.set_index(df.columns[0])

        def get_val(row_name, yr):
            yr_cols = [c for c in df.columns if str(yr) in str(c)]
            if not yr_cols: return None
            try: return float(str(df.loc[row_name, yr_cols[0]]).replace(",",""))
            except: return None

        revenue_cur  = get_val("营业总收入", year)
        revenue_prev = get_val("营业总收入", int(year)-1)
        profit_cur   = get_val("净利润", year)
        profit_prev  = get_val("净利润", int(year)-1)
        gross_profit = get_val("毛利润", year) or get_val("营业利润", year)

        if revenue_cur and revenue_prev and revenue_prev != 0:
            data["revenue_growth"] = (revenue_cur - revenue_prev) / abs(revenue_prev) * 100
        if profit_cur and profit_prev and profit_prev != 0:
            data["profit_growth"] = (profit_cur - profit_prev) / abs(profit_prev) * 100
        if revenue_cur and gross_profit:
            data["gross_margin"] = gross_profit / revenue_cur * 100
        if revenue_cur and profit_cur:
            data["net_margin"] = profit_cur / revenue_cur * 100

        # 毛利额增速
        gross_prev = get_val("毛利润", int(year)-1)
        if gross_profit and gross_prev and gross_prev != 0:
            data["gross_profit_yoy"] = (gross_profit - gross_prev) / abs(gross_prev) * 100

    except Exception as e:
        pass

    # ── 资产负债表 ──
    try:
        df_bs = ak.stock_balance_sheet_by_yearly_em(symbol=code)
        df_bs.columns = [str(c).strip() for c in df_bs.columns]
        df_bs = df_bs.set_index(df_bs.columns[0])

        def get_bs(row_name, yr):
            yr_cols = [c for c in df_bs.columns if str(yr) in str(c)]
            if not yr_cols: return None
            try: return float(str(df_bs.loc[row_name, yr_cols[0]]).replace(",",""))
            except: return None

        total_assets    = get_bs("资产总计", year)
        total_liab      = get_bs("负债合计", year)
        current_assets  = get_bs("流动资产合计", year)
        current_liab    = get_bs("流动负债合计", year)
        receivables     = get_bs("应收账款", year)

        if total_assets and total_liab:
            data["debt_ratio"] = total_liab / total_assets * 100
        if current_assets and current_liab and current_liab != 0:
            data["current_ratio"] = current_assets / current_liab
        if receivables and revenue_cur if "revenue_growth" in data else False:
            pass  # 后面用营收再算

    except Exception:
        pass

    # ── 现金流量表 ──
    try:
        df_cf = ak.stock_cash_flow_sheet_by_yearly_em(symbol=code)
        df_cf.columns = [str(c).strip() for c in df_cf.columns]
        df_cf = df_cf.set_index(df_cf.columns[0])

        def get_cf(row_name, yr):
            yr_cols = [c for c in df_cf.columns if str(yr) in str(c)]
            if not yr_cols: return None
            try: return float(str(df_cf.loc[row_name, yr_cols[0]]).replace(",",""))
            except: return None

        cfo   = get_cf("经营活动产生的现金流量净额", year)
        capex = get_cf("购建固定资产、无形资产和其他长期资产支付的现金", year)
        profit_cur2 = data.get("net_margin")  # 用已取到的

        if cfo:
            # 经营现金/净利润
            try:
                df_p = ak.stock_profit_sheet_by_yearly_em(symbol=code)
                df_p.columns = [str(c).strip() for c in df_p.columns]
                df_p = df_p.set_index(df_p.columns[0])
                yr_cols = [c for c in df_p.columns if str(year) in str(c)]
                if yr_cols:
                    np_val = float(str(df_p.loc["净利润", yr_cols[0]]).replace(",",""))
                    if np_val and np_val != 0:
                        data["cfo_to_profit"] = cfo / np_val
            except Exception:
                pass

    except Exception:
        pass

    # ── ROE（用主要指标接口）──
    try:
        df_key = ak.stock_financial_abstract_ths(symbol=code, indicator="按年度")
        yr_row = df_key[df_key["报告期"].astype(str).str.contains(str(year))]
        if not yr_row.empty:
            r = yr_row.iloc[0]
            def sf(col):
                try: return float(str(r[col]).replace("%","").replace(",",""))
                except: return None
            data.setdefault("roe",        sf("净资产收益率"))
            data.setdefault("gross_margin", sf("毛利率"))
            data.setdefault("net_margin", sf("净利率"))
            data.setdefault("debt_ratio", sf("资产负债率"))
            data.setdefault("revenue_growth", sf("营业总收入同比增长率"))
            data.setdefault("profit_growth",  sf("归母净利润同比增长率"))
            data.setdefault("eps_growth",     sf("每股收益同比增长率"))
    except Exception:
        pass

    return data


def fetch_company_data(code: str, year: str) -> dict:
    """汇总一家公司所有指标"""
    console.print(f"  📡 {code} 数据获取中...", style="dim")
    data = {"code": code}
    data["name"]  = get_stock_name(code)
    data.update(fetch_realtime_metrics(code))
    data.update(fetch_financial_metrics(code, year))
    return data


def rank_mark(values: list, idx: int, higher_is_better: bool) -> str:
    """给当前值打排名标记"""
    valid = [(i, v) for i, v in enumerate(values) if v is not None]
    if len(valid) < 2:
        return ""
    sorted_vals = sorted(valid, key=lambda x: x[1], reverse=higher_is_better)
    rank = next((r for r, (i, _) in enumerate(sorted_vals) if i == idx), None)
    if rank is None:
        return ""
    if rank == 0:
        return " 🥇"
    elif rank == 1:
        return " 🥈"
    elif rank == len(valid) - 1:
        return " 🔴"
    return ""


def format_val(v, fmt, unit) -> str:
    if v is None:
        return "N/A"
    try:
        s = f"{v:{fmt}}"
        return f"{s}{unit}"
    except Exception:
        return str(v)


def print_section(companies: list, section_key: str):
    """打印一个分组的对比表"""
    section  = METRICS[section_key]
    fields   = section["fields"]

    table = Table(
        title=section["title"],
        box=box.ROUNDED,
        show_header=True,
        header_style="bold cyan",
        title_style="bold yellow",
        min_width=80,
    )

    table.add_column("指标", style="bold white", min_width=14)
    for c in companies:
        table.add_column(f"{c['name']}\n({c['code']})",
                         justify="right", min_width=12)

    for (field, label, unit, higher, fmt) in fields:
        vals = [c.get(field) for c in companies]
        row  = [f"{label} ({unit})"]
        for i, v in enumerate(vals):
            mark  = rank_mark(vals, i, higher) if higher is not None else ""
            style = ""
            if mark == " 🥇":  style = "bold green"
            elif mark == " 🔴": style = "bold red"
            cell = Text(format_val(v, fmt, "") + mark, style=style)
            row.append(cell)
        table.add_row(*row)

    console.print(table)
    console.print()


def export_excel(companies: list, output_path: str, year: str):
    """导出 Excel 对比报告"""
    rows = []
    for section_key, section in METRICS.items():
        for (field, label, unit, higher, fmt) in section["fields"]:
            row = {"分类": section["title"], "指标": f"{label}（{unit}）"}
            for c in companies:
                v = c.get(field)
                row[f"{c['name']}({c['code']})"] = (
                    float(f"{v:{fmt}}") if v is not None else None
                )
            rows.append(row)

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="同业对比", index=False)

        ws = writer.sheets["同业对比"]
        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 25)

        # 标题行加色
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 数字列右对齐
        for row in ws.iter_rows(min_row=2):
            for cell in row[2:]:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

    console.print(f"[green]📊 Excel 报告已导出：{output_path}[/green]")


def print_summary(companies: list):
    """综合评分小结"""
    console.rule("[bold yellow]🏅 综合评分小结[/bold yellow]")

    scores = {c["code"]: 0 for c in companies}
    total  = 0

    for section_key, section in METRICS.items():
        for (field, label, unit, higher, fmt) in section["fields"]:
            if higher is None:
                continue
            vals = [(i, c.get(field)) for i, c in enumerate(companies)]
            valid = [(i, v) for i, v in vals if v is not None]
            if len(valid) < 2:
                continue
            sorted_vals = sorted(valid, key=lambda x: x[1], reverse=higher)
            for rank, (i, _) in enumerate(sorted_vals):
                pts = len(valid) - rank
                scores[companies[i]["code"]] += pts
            total += len(valid)

    sorted_companies = sorted(companies, key=lambda c: scores[c["code"]], reverse=True)
    medals = ["🥇", "🥈", "🥉"] + ["  "] * 10

    table = Table(box=box.SIMPLE, show_header=True, header_style="bold cyan")
    table.add_column("排名", justify="center", width=6)
    table.add_column("公司")
    table.add_column("代码")
    table.add_column("综合得分", justify="right")
    table.add_column("满分", justify="right")

    for i, c in enumerate(sorted_companies):
        score = scores[c["code"]]
        table.add_row(
            medals[i],
            Text(c["name"], style="bold" if i == 0 else ""),
            c["code"],
            Text(str(score), style="green bold" if i == 0 else ""),
            str(total),
        )

    console.print(table)
    console.print(
        "[dim]注：综合得分基于各项指标相对排名加总，数值越高表示在同业中整体表现越优。"
        "需结合行业特性和个人投资逻辑综合判断。[/dim]\n"
    )


def main():
    parser = argparse.ArgumentParser(description="同业财务对比")
    parser.add_argument("--stocks", nargs="+", required=True,
                        help="2~6只股票代码，如：600519 000858 000568")
    parser.add_argument("--year",   default=str(datetime.date.today().year - 1),
                        help="财报年份（默认上一年）")
    parser.add_argument("--output", default=None,
                        help="导出Excel路径，如 report.xlsx")
    args = parser.parse_args()

    if len(args.stocks) < 2:
        console.print("[red]至少输入2只股票代码进行对比。[/red]")
        sys.exit(1)
    if len(args.stocks) > 6:
        console.print("[yellow]建议最多对比6只，过多会影响可读性。[/yellow]")

    console.rule(f"[bold cyan]同业财务对比  {args.year}年报[/bold cyan]")
    console.print(f"对比标的：{' | '.join(args.stocks)}\n", style="dim")

    companies = []
    for code in args.stocks:
        try:
            c = fetch_company_data(code, args.year)
            companies.append(c)
        except Exception as e:
            console.print(f"[red]⚠️  {code} 数据获取失败：{e}[/red]")

    if len(companies) < 2:
        console.print("[red]有效数据不足2家，无法对比。[/red]")
        sys.exit(1)

    console.print()
    for section_key in METRICS:
        print_section(companies, section_key)

    print_summary(companies)

    if args.output:
        export_excel(companies, args.output, args.year)
    else:
        # 默认导出到 outputs
        out = f"/mnt/user-data/outputs/peer_compare_{args.year}_{'_'.join(args.stocks[:3])}.xlsx"
        export_excel(companies, out, args.year)


if __name__ == "__main__":
    main()
